import discord
from discord.ext import commands
from datetime import datetime
from random import choice
from openpyxl import load_workbook
TOKEN = ""
bot = commands.Bot(command_prefix="!", help_command=None)
    
@bot.command()
async def randomize(ctx):
    """[Command to generate new teams reorganization]

    Args:
        ctx ([str]): [command info]
    """    
    found = 0
    #Initializes a Discord Embed
    embedVar = discord.Embed(title="Teams",description=f"Date: {datetime.now().day}/{datetime.now().month}" , color=0x32a6a8)
    #Defined the possible teams
    lista = [[],[],[],[],[],[]]
    #Iterates over every possible team group
    for current in lista:
        while True:
            random = choice(range(24))
            #Every team should be conformed of 4 people only
            if len(current) != 4:
                for x in lista:
                    #Avoid repetitions
                    if random in x:
                        found = 1
                        break
                #Set a new student number
                if found == 0:
                    current.append(random)
                found = 0
            else:
                break
    #Looking into the excel information
    wb = load_workbook("names_list.xlsx")
    #Setting a new worksheet
    ws = wb[wb.sheetnames[0]]
    #The actual generation of the Embed Content
    for index, teams in enumerate(lista):
        names = ""
        for student in teams:
            names += ws[f"C{8+student}"].value + " " +ws[f"B{8+student}"].value +"\n"
        embedVar.add_field(name=f"Team #{index + 1}", value=names, inline=False)
    dev = await bot.fetch_user(257009761571176458)
    embedVar.set_footer(text="Developer: Snow!!", icon_url=dev.avatar_url)
    await ctx.send(embed=embedVar)
            
        





@bot.event
async def on_message(message):
    await bot.process_commands(message)

@bot.event
async def on_ready():
    print("Ready")


bot.run(TOKEN)
